import csv
import openpyxl

def convertMatrix(inputfile,outputFile):

    wb = openpyxl.Workbook()
    ws = wb.active

    with open(inputfile) as csvMatrix:
        reader = csv.reader(csvMatrix, delimiter=',')
        count_row = 1
        for row in reader:
            count_column = 1
            for cell in row:
                if cell == 'nb':
                    ws.cell(row=count_row, column=count_column).value = ''
                else:
                    try:
                        ws.cell(row=count_row, column=count_column).value = 1/float(cell)
                    except:
                        ws.cell(row=count_row, column=count_column).value = cell
                count_column += 1
            count_row +=1

    wb.save(outputFile)







if __name__ == "__main__":
    convertMatrix('Lift_Matrix.csv','dissimilarity_matrix.xlsx')