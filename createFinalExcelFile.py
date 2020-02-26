import openpyxl
import csv

def extractionMatrix(candidates, issues):
    rowColMatch = []
    curr_row = 2
    for i in candidates:
        curr_column = 2
        for j in issues:
            rowColMatch.append((curr_column+len(candidates),curr_row))
            curr_column += 1
        curr_row += 1
    return rowColMatch

def extractSentimentAvg(candidate,issue):
    wb = openpyxl.load_workbook('./SentimentFiles/'+candidate+issue+'.xlsx', data_only=True)
    ws = wb.active
    if ws.max_row == 1:
        return (0,0)
    score = ws.cell(row=ws.max_row-1, column=5).value
    stDev = ws.cell(row=ws.max_row, column=5).value

    return (score, stDev)

def createExcelSheet(candidates,issues, outputFile, liftMatrix):

    liftValWb = openpyxl.load_workbook(liftMatrix)
    liftValWs = liftValWb.active
    liftPairs = extractionMatrix(candidates, issues)

    wb = openpyxl.Workbook()
    ws = wb.active
    row_count = 1
    ws.cell(row=row_count, column=1).value = "Candidate_Issue Pair"
    ws.cell(row=row_count, column=2).value = "Lift_Value"
    ws.cell(row=row_count, column=3).value = "Sentiment Score"
    ws.cell(row=row_count, column=4).value = "Sentiment St-Dev"
    row_count = 2

    for candidate in candidates:
        for issue in issues:
            sentimentScore = extractSentimentAvg(candidate, issue)
            ws.cell(row=row_count, column=1).value = candidate+'_'+issue
            ws.cell(row=row_count, column=2).value = 1/liftValWs.cell(row=liftPairs[row_count-2][1], column=liftPairs[row_count-2][0]).value
            ws.cell(row=row_count, column=3).value = sentimentScore[0]
            ws.cell(row=row_count, column=4).value = sentimentScore[1]
            row_count +=1

    wb.save(outputFile)

if __name__ == "__main__":

    candidates = ['bernie','bloomberg']
    issues = ['economy','climate','immigration']
    liftMatrix = extractionMatrix(candidates, issues)
    createExcelSheet(candidates, issues, 'finalTable.xlsx','OutputFiles/dissimilarity_matrix.xlsx')
