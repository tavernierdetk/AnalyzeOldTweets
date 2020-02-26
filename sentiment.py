#!/usr/bin/python2.7
# the input filename is limit_post.csv in line 33, change it as needed
# the output file is sentiment_data.xlsx, change it when running the script multiple times

import pandas as pd
import numpy as np
import nltk
import openpyxl
import os
import statistics


def get_sentiment(rating_data):
    """
    https: // github.com / cjhutto / vaderSentiment
    :return:
    """
    from nltk.sentiment.vader import SentimentIntensityAnalyzer
    sid = SentimentIntensityAnalyzer()
    rating_data['sent_neg'] = -10
    rating_data['sent_neu'] = -10
    rating_data['sent_pos'] = -10
    rating_data['sent_compound'] = -10
    for i in range(len(rating_data)):
        sentence = rating_data['Sentences'][i]
#         print sentence
        ss = sid.polarity_scores(sentence)
        rating_data.iloc[i, 1] = float(ss['neg'])
        rating_data.iloc[i, 2] = ss['neu']
        rating_data.iloc[i, 3] = ss['pos']
        rating_data.iloc[i, 4] = ss['compound']
    return rating_data

def runSentimentAnalysis(inputFile):
    fileToRun = inputFile+'.csv'
    rating_data = pd.read_csv(fileToRun, encoding = 'latin1')
    rating_data = rating_data.rename(columns={ rating_data.columns[0]: "Sentences" })
    sentiment_data = get_sentiment(rating_data)
    outputFile = inputFile+'.xlsx'
    sentiment_data.to_excel(outputFile, index = False)
    print("Written to: " + outputFile)

def computeAverage(inputFile):
    fileToLad = inputFile + '.xlsx'
    wb = openpyxl.load_workbook(fileToLad)
    ws = wb.active
    lastRow = ws.max_row+1
    runningTotal = 0
    totalsArray = []
    for i in range(2, ws.max_row):
        runningTotal += float(ws.cell(column=5, row=i).value)
        totalsArray.append(float(ws.cell(column=5, row=i).value))
    if ws.max_row == 1:
        return
    ws.cell(row=lastRow, column=5).value = runningTotal/(ws.max_row-1)
    ws.cell(row=lastRow+1, column=5).value = statistics.stdev(totalsArray)


    wb.save(fileToLad)

if __name__ == "__main__":
    runSentimentAnalysis('./SentimentFiles/'+'bernieclimate')
    computeAverage('./SentimentFiles/'+'bernieclimate')
